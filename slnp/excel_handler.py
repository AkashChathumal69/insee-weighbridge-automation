import pandas as pd
import json
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

EXPORT_PATH = "./Database/exports"

# Ensure export directory exists
os.makedirs(EXPORT_PATH, exist_ok=True)


class ExcelHandler:
    """Handle Excel import/export operations"""
    
    @staticmethod
    def export_to_excel(processes):
        """Export process data to Excel"""
        try:
            if not processes:
                return None
            
            # Flatten the nested data structure for Excel
            rows = []
            for process in processes:
                # Extract wait_in data
                wait_in = process.get("waitIn", {})
                delivery_table = wait_in.get("deliveryTable", [])
                
                # Create a base row
                base_row = {
                    "Token Number": process.get("tokenNumber", ""),
                    "Vehicle Number": process.get("vehicleNumber", ""),
                    "Date": process.get("date", ""),
                    "Arrival Time": process.get("arrivalTime", ""),
                    "Status": process.get("status", ""),
                    "Driver Name": wait_in.get("driverName", ""),
                    "Driver Phone": wait_in.get("driverPhone", ""),
                    "Driver Town": wait_in.get("driverTown", ""),
                    "Driver License": wait_in.get("driverLicense", ""),
                    "Driver Alcohol Test": wait_in.get("driverAlcoholTest", ""),
                    "Helper Name": wait_in.get("helperName", ""),
                    "Helper Identity": wait_in.get("helperIdentity", ""),
                    "Helper Phone": wait_in.get("helperPhone", ""),
                    "Helper Town": wait_in.get("helperTown", ""),
                    "Helper Alcohol Test": wait_in.get("helperAlcoholTest", ""),
                    "Vehicle Insurance": wait_in.get("vehicleInsurance", False),
                    "Driver PPE Number": wait_in.get("driverPPENumber", ""),
                    "Helper PPE Number": wait_in.get("helperPPENumber", ""),
                }
                
                # Add delivery table data
                for i, delivery in enumerate(delivery_table):
                    base_row[f"{delivery.get('brand', 'Brand')} Requested"] = delivery.get("requestedBag", 0)
                    base_row[f"{delivery.get('brand', 'Brand')} Delivered"] = delivery.get("deliveryBag", 0)
                
                # Add wait_out data if exists
                wait_out = process.get("waitOut")
                if wait_out:
                    base_row["Departure Time"] = wait_out.get("departureTime", "")
                    base_row["Total Issue"] = wait_out.get("totalIssue", "")
                    base_row["Notes"] = wait_out.get("notes", "")
                
                rows.append(base_row)
            
            # Create DataFrame
            df = pd.DataFrame(rows)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"process_data_{timestamp}.xlsx"
            filepath = os.path.join(EXPORT_PATH, filename)
            
            # Export to Excel
            df.to_excel(filepath, sheet_name="Process Data", index=False)
            
            # Format the Excel file
            ExcelHandler._format_excel(filepath)
            
            return filepath
        
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None
    
    @staticmethod
    def _format_excel(filepath):
        """Format Excel file with colors and styles"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
        except Exception as e:
            print(f"Error formatting Excel: {e}")
    
    @staticmethod
    def import_from_excel(filepath):
        """Import data from Excel file"""
        try:
            df = pd.read_excel(filepath)
            
            # Convert DataFrame to list of dictionaries
            data = df.to_dict('records')
            
            return data
        
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            return None
    
    @staticmethod
    def get_latest_export():
        """Get the path to the latest export file"""
        try:
            if not os.path.exists(EXPORT_PATH):
                return None
            
            files = os.listdir(EXPORT_PATH)
            if not files:
                return None
            
            # Get the most recently created file
            latest = max(files, key=lambda f: os.path.getctime(os.path.join(EXPORT_PATH, f)))
            return os.path.join(EXPORT_PATH, latest)
        
        except Exception as e:
            print(f"Error getting latest export: {e}")
            return None
    
    @staticmethod
    def list_exports():
        """List all exported files"""
        try:
            if not os.path.exists(EXPORT_PATH):
                return []
            
            files = os.listdir(EXPORT_PATH)
            return sorted(files, key=lambda f: os.path.getctime(os.path.join(EXPORT_PATH, f)), reverse=True)
        
        except Exception as e:
            print(f"Error listing exports: {e}")
            return []
